import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# STEP 1: Read Excel
filename = "TimeTable.xlsx"
sheet_name = "Time Table"
df_raw = pd.read_excel(filename, sheet_name=sheet_name, header=None, skiprows=3)

# STEP 2: Input course list
input_str = input("Enter course codes to include (comma-separated): ")
selected_courses = [code.strip().upper() for code in input_str.split(",")]

# STEP 3: Prepare time slots
time_slots = df_raw.iloc[0, 1:].tolist()
df = df_raw.iloc[1:].copy()
df.columns = ['Day'] + time_slots

# STEP 4: Group entries by day and slot
grouped_data = defaultdict(lambda: {slot: [] for slot in time_slots})

current_day = None
for i, row in df.iterrows():
    day = row['Day']
    if pd.notna(day):
        current_day = day.strip()
    if current_day is None:
        continue
    for slot in time_slots:
        cell = row[slot]
        if pd.isna(cell):
            continue
        entries = str(cell).split('\n') if '\n' in str(cell) else [cell]
        for entry in entries:
            for course in selected_courses:
                if course in entry:
                    parts = entry.split('(')
                    code = parts[0].strip()
                    room = '(' + parts[-1] if len(parts) > 1 else ''
                    formatted = f"{code}<br><small>{room}</small>"
                    grouped_data[current_day][slot].append(formatted)



wb = Workbook()
ws = wb.active
ws.title = "Custom Timetable"

# Styles
header_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")

# Write headers
ws.cell(row=1, column=1, value="Day").font = header_font
ws.cell(row=1, column=1).alignment = center_align
ws.cell(row=1, column=1).fill = header_fill
ws.cell(row=1, column=1).border = thin_border

for idx, slot in enumerate(time_slots):
    cell = ws.cell(row=1, column=2 + idx, value=slot)
    cell.font = header_font
    cell.alignment = center_align
    cell.fill = header_fill
    cell.border = thin_border

# Write data
row_num = 2
for day, slots in grouped_data.items():
    ws.cell(row=row_num, column=1, value=day).alignment = center_align
    ws.cell(row=row_num, column=1).border = thin_border
    for col_num, slot in enumerate(time_slots, start=2):
        entries = slots[slot]
        content = "\n".join(entry.replace("<br>", "\n").replace("<small>", "").replace("</small>", "") for entry in entries)
        cell = ws.cell(row=row_num, column=col_num, value=content)
        cell.alignment = center_align
        cell.border = thin_border
    row_num += 1

wb.save("my_timetable.xlsx")
print("\n✅ Timetable saved as: my_timetable.xlsx")
